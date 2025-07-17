# GOBIERNO/generador_maestro.py

import os
import pandas as pd
from pathlib import Path
import threading
import re
import locale
import math
from datetime import datetime
import sys
from openpyxl import load_workbook, Workbook # Necesario para guardar_bd_maestra_unificada
from openpyxl.utils.dataframe import dataframe_to_rows # Necesario para guardar_bd_maestra_unificada
from openpyxl.utils.exceptions import InvalidFileException # Para manejo de errores en carga
import threading
import unicodedata
from docx import Document
from docx.shared import Pt, Inches  
import time  # Para las pausas
import traceback # Para mostrar errores inesperados detallados
import paramiko 
import shutil
import copy

import GeneradorPredial_logica
import GeneradorMultas_logica

# --- CONSTANTES GLOBALES DEL SCRIPT INTEGRADOR ---
SCRIPT_BASE_PATH = Path(__file__).resolve().parent  # Directorio GOBIERNO/
KEYWORD_MULTAS = "MULTAS"
KEYWORD_PREDIAL = "PREDIAL"

# Al inicio de generador_maestro.py, DESPUÉS de definir SCRIPT_BASE_PATH
# Asegúrate que GeneradorFER_logica.py está en el mismo directorio o en PYTHONPATH
try:
    import GeneradorPredial_logica # Asumiendo que el archivo se llama así
except ImportError as e:
    print(f"ERROR CRITICO: No se pudo importar GeneradorPredial_logica.py. Asegúrate que está en {SCRIPT_BASE_PATH}.")
    print(e)
    sys.exit(1)

# --- Definiciones de Columnas y Tipos para BDs Maestras ---

# --- Definiciones para PREDIAL ---

# EXPLICACIÓN: Se define la variable ANTES de ser utilizada y con el nombre estandarizado.
COL_BASE_DATOS_ESCANEO_PREDIAL = "BASE DE DATOS ESCANEO" 

# MODIFICADO: Nueva estructura de columnas para la BD Maestra de Predial
COLUMNAS_BD_MAESTRA_PREDIAL = [
    "ID", "EXPEDIENTE", "NOMBRE CONTRIBUYENTE", "DIRECCION", "COLONIA", 
    "BIMESTRE", "AÑOS", "CLAVE DE EXTENCION", "TIPO", "MONTO", "ESTADO", 
    "HOJAS POR DOCUMENTO", "FECHA IMPRESION", COL_BASE_DATOS_ESCANEO_PREDIAL, 
    "Ruta PDF Escaneado", "Ruta PDF Generado", "MOVIMIENTO"
]

# MODIFICADO: Tipos de datos para la nueva estructura (MONTO es ahora float64)
TIPOS_BD_MAESTRA_PREDIAL = {
    "ID": 'str',
    "EXPEDIENTE": 'str',
    "NOMBRE CONTRIBUYENTE": 'str',
    "DIRECCION": 'str',
    "COLONIA": 'str',
    "BIMESTRE": 'str',
    "AÑOS": 'str',
    "CLAVE DE EXTENCION": 'str',
    "TIPO": 'str',
    "MONTO": 'float64', # <-- TIPO NUMÉRICO
    "ESTADO": 'str',
    "HOJAS POR DOCUMENTO": 'Int64',
    "FECHA IMPRESION": 'str',
    COL_BASE_DATOS_ESCANEO_PREDIAL: 'str',
    "Ruta PDF Escaneado": 'str',
    "Ruta PDF Generado": 'str',
    "MOVIMIENTO": 'str'
}

# La columna de búsqueda en el archivo de datos sigue siendo EXPEDIENTE
COL_ID_PREDIAL = "EXPEDIENTE" 
COL_ESTADO_PREDIAL_BD = "ESTADO"
COL_NOMBRE_CSV_PREDIAL = "NOMBRE"
COL_COLONIA_CSV_PREDIAL = "COLONIA"

# --- Definiciones para MULTAS ---

# EXPLICACIÓN: Se define la variable ANTES de ser utilizada.
COL_CONTEO_REGISTROS_MULTAS = "REGISTROS EN BD GENERACION"
COL_BASE_DATOS_ESCANEO_MULTAS = "BASE DE DATOS ESCANEO"

# MODIFICADO: Nueva estructura de columnas para la BD Maestra de Multas, ya SIN las columnas eliminadas.
COLUMNAS_BD_MAESTRA_MULTAS = [
    "ID",
    "OFICIO",
    "NOMBRE CONTRIBUYENTE",
    "DIRECCION",
    "CP",
    "MONTO",
    "ESTADO",
    "HOJAS POR DOCUMENTO",
    "REGISTROS EN BD GENERACION", # Para contar las filas del oficio
    "FECHA IMPRESION",
    "BASE DE DATOS ESCANEO", # Nombre del archivo Excel del escaneo
    "Ruta PDF Escaneado",
    "Ruta PDF Generado",
    "MOVIMIENTO"
]

# MODIFICADO: Nuevos tipos de datos, con MONTO como número.
TIPOS_BD_MAESTRA_MULTAS = {
    "ID": 'str',
    "OFICIO": 'str',
    "NOMBRE CONTRIBUYENTE": 'str',
    "DIRECCION": 'str',
    "CP": 'str',
    "MONTO": 'float64', # TIPO NUMÉRICO PARA SUMAS CORRECTAS
    "ESTADO": 'str',
    "HOJAS POR DOCUMENTO": 'Int64',
    "REGISTROS EN BD GENERACION": 'Int64', # TIPO NUMÉRICO
    "FECHA IMPRESION": 'str',
    "BASE DE DATOS ESCANEO": 'str',
    "Ruta PDF Escaneado": 'str',
    "Ruta PDF Generado": 'str',
    "MOVIMIENTO": 'str'
}

# La columna de búsqueda en la BD Maestra y en el CSV de datos sigue siendo OFICIO
COL_ID_MULTAS_BD = "OFICIO"
COL_ID_MULTAS_CSV = "OFICIO"
COL_ESTADO_MULTAS_BD = "ESTADO"
COL_NOMBRE_CSV_MULTAS_BASE = "NOMBRE"
COL_APATERNO_CSV_MULTAS_BASE = "APELLIDO PATERNO"
COL_AMATERNO_CSV_MULTAS_BASE = "APELLIDO MATERNO"
COL_CP_CSV_MULTAS = "CP"


# --- ESTADOS GLOBALES UNIFICADOS ---
ESTADO_PENDIENTE = "PENDIENTE"
ESTADO_GEN_COMPLETO = "GENERADO COMPLETO"
ESTADO_GEN_ULTIMA = "GENERADO ULTIMA HOJA"
ESTADO_GEN_RESTO = "GENERADO RESTO (SIN ULTIMA)"
ESTADO_IMP_ULTIMA = "ULTIMA PAG IMPRESA"
ESTADO_IMP_COMPLETO = "IMPRESION COMPLETADA"
ESTADO_ERROR_GENERACION = "ERROR GENERACION"
ESTADO_PDF_NO_ENCONTRADO = "PDF NO ENCONTRADO"
ESTADO_NO_GENERADO_MONTO_BAJO = "NO GENERADO (MENOR A 179)"


MODO_GENERACION_COMPLETO = "COMPLETO"
MODO_GENERACION_ULTIMA = "ULTIMA"
MODO_GENERACION_RESTO = "RESTO"
MODO_GENERACION_ESPECIFICOS = "ESPECIFICOS"
MODO_ACCION_ESCANEAR_MULTAS = "ESCANEAR_MULTAS"
MODO_ACCION_ESCANEAR_PREDIAL = "ESCANEAR_PREDIAL" # Nueva constante para Predial
MODO_ACCION_REPORTE_DESPACHOS_MULTAS = "REPORTE_DESPACHOS_MULTAS" # << NUEVA CONSTANTE




excel_lock = threading.Lock()


def cargar_bd_maestra_unificada(mode_config):
    """
    Carga la Base de Datos Maestra desde el archivo Excel especificado en mode_config.
    Si el archivo no existe, o la hoja 'BD_Maestra' no existe,
    crea un DataFrame vacío con la estructura definida.
    Asegura que todas las columnas definidas en mode_config["db_master_columns"] existan
    y tengan los tipos de datos definidos en mode_config["db_master_types"].
    Las columnas ID (definidas por mode_config["col_expediente"]) se fuerzan a string.
    *** VERSIÓN MODIFICADA PARA PRESERVAR COLUMNAS EXTRA COMO 'MOVIMIENTO' ***
    """
    ruta_excel_bd = Path(mode_config["master_db_file_path"])
    sheet_name_bd = "BD_Maestra"
    defined_columns = mode_config["db_master_columns"]
    defined_types = mode_config["db_master_types"]
    id_column_name = mode_config["col_expediente"]
    global excel_lock

    with excel_lock:
        if ruta_excel_bd.exists():
            try:
                print(f"    - Cargando BD Maestra para '{mode_config['mode_name']}' desde: {ruta_excel_bd}")
                dtype_read = {col: str for col in defined_columns}

                df = pd.read_excel(
                    ruta_excel_bd,
                    sheet_name=sheet_name_bd,
                    # No especificamos dtype aquí para leer todas las columnas tal como vienen
                    keep_default_na=False,
                    na_filter=False
                )
                
                # --- INICIO DEL CAMBIO ---
                # Guardar los nombres de las columnas originales del Excel
                original_excel_columns = [str(col).strip() for col in df.columns]
                df.columns = original_excel_columns
                # --- FIN DEL CAMBIO ---

                # Forzar la columna ID a ser un string limpio ANTES de cualquier otra cosa
                if id_column_name in df.columns:
                     df[id_column_name] = df[id_column_name].astype(str).str.strip()
                     # Si la columna puede venir como float (ej. "12345.0"), esto lo corrige a "12345"
                     df[id_column_name] = df[id_column_name].str.replace(r'\.0$', '', regex=True)

                for col_name in defined_columns:
                    expected_dtype_str = defined_types.get(col_name)
                    if not expected_dtype_str:
                        print(f"      - (!) Advertencia: Tipo no definido para columna '{col_name}' en mode_config. Se asumirá 'str'.")
                        expected_dtype_str = 'str'

                    if col_name not in df.columns:
                        print(f"      -> Columna '{col_name}' no encontrada en Excel. Añadiendo...")
                        if expected_dtype_str == 'Int64':
                            df[col_name] = pd.NA
                        else:
                            df[col_name] = ""

                    # --- Lógica de conversión de tipos (sin cambios) ---
                    try:
                        if expected_dtype_str == 'str':
                            df[col_name] = df[col_name].fillna("").astype(str)
                        elif expected_dtype_str == 'Int64':
                            df[col_name] = pd.to_numeric(df[col_name], errors='coerce').astype('Int64')
                        else:
                            if expected_dtype_str in ['float64', 'float32', 'int32']:
                                df[col_name] = pd.to_numeric(df[col_name].replace('', pd.NA), errors='coerce').astype(expected_dtype_str)
                            else:
                                df[col_name] = df[col_name].astype(expected_dtype_str)
                    except Exception as e_type_conv:
                        print(f"      - (!) Advertencia: No se pudo convertir columna '{col_name}'. Se usará string. Error: {e_type_conv}")
                        df[col_name] = df[col_name].astype(str).fillna("")

                # --- INICIO DEL CAMBIO ---
                # Construir la lista final de columnas, preservando las extras como 'MOVIMIENTO'
                final_columns_to_keep = list(defined_columns)
                for excel_col in original_excel_columns:
                    if excel_col not in final_columns_to_keep:
                        final_columns_to_keep.append(excel_col)
                
                if id_column_name in df.columns:
                     df[id_column_name] = df[id_column_name].fillna("").astype(str)

                # Reordenar para que las columnas definidas por el script estén primero, y las extras al final.
                df = df.reindex(columns=final_columns_to_keep)
                # --- FIN DEL CAMBIO ---
                
                print(f"    - BD Maestra para '{mode_config['mode_name']}' cargada. {len(df)} registros. Columnas detectadas: {df.columns.tolist()}")
                return df
            except InvalidFileException:
                print(f"    - (!) ERROR CRITICO: Archivo BD Maestra '{ruta_excel_bd}' está corrupto.")
                sys.exit(1)
            except Exception as e_general_load:
                print(f"    - (!) ADVERTENCIA GENERAL: Error inesperado al cargar BD Maestra '{ruta_excel_bd}'. Se creará/usará una BD vacía. Error: {e_general_load}")
        else:
            print(f"    - Archivo BD Maestra '{ruta_excel_bd}' no encontrado. Se creará uno nuevo (sin columnas extra).")

        # El código para crear un DataFrame vacío no cambia, ya que no tendría la columna 'MOVIMIENTO' de todas formas.
        df_vacio = pd.DataFrame(columns=defined_columns)
        for col, dtype_val in defined_types.items():
            if dtype_val == 'str':
                df_vacio[col] = pd.Series(dtype='object').fillna("")
            else:
                df_vacio[col] = pd.Series(dtype=dtype_val)
        df_vacio = df_vacio.astype(defined_types)
        return df_vacio
    
def obtener_siguiente_letra_lote(letra_actual):
    """
    Obtiene la siguiente letra del alfabeto para el lote.
    Ej: 'A' -> 'B', 'Z' -> 'AA', 'AZ' -> 'BA'.
    """
    if not letra_actual or not letra_actual.isalpha():
        return 'A'
    
    letra_actual = letra_actual.upper()
    if letra_actual == 'Z' * len(letra_actual):
        return 'A' * (len(letra_actual) + 1)
    
    letras = list(letra_actual)
    i = len(letras) - 1
    while i >= 0:
        if letras[i] == 'Z':
            letras[i] = 'A'
            i -= 1
        else:
            letras[i] = chr(ord(letras[i]) + 1)
            break
    return "".join(letras)

def obtener_ultimo_id(df_bd_maestra):
    """
    Lee el DataFrame de la BD Maestra y devuelve el último ID de lote.
    Devuelve (letra, numero) ej: ('D', 20) o (None, 0) si no hay IDs.
    """
    if df_bd_maestra.empty or "ID" not in df_bd_maestra.columns:
        return None, 0

    # Filtra IDs que sigan el patrón LETRA(S)-NUMERO para evitar errores
    ids_validos = df_bd_maestra[df_bd_maestra['ID'].astype(str).str.match(r'^[A-Z]+-\d+', na=False, case=False)].copy()
    if ids_validos.empty:
        return None, 0

    # Para encontrar el último ID real, debemos ordenar correctamente
    ids_validos['letra_lote'] = ids_validos['ID'].str.split('-').str[0].str.upper()
    ids_validos['numero_lote'] = pd.to_numeric(ids_validos['ID'].str.split('-').str[1])
    
    # Ordena primero por la longitud de la letra (para que 'Z' venga antes de 'AA'), luego por la letra y finalmente por el número
    ids_validos['len_letra'] = ids_validos['letra_lote'].str.len()
    ids_validos = ids_validos.sort_values(by=['len_letra', 'letra_lote', 'numero_lote'])
    
    ultimo_id_str = ids_validos['ID'].iloc[-1]
    
    try:
        partes = ultimo_id_str.split('-')
        letra = partes[0].upper()
        numero = int(partes[1])
        return letra, numero
    except (IndexError, ValueError):
        return None, 0
    
def obtener_ultimo_id_de_lote_especifico(df_bd_maestra, letra_lote):
    """
    Encuentra el número más alto usado en un lote específico (ej. 'A').
    Devuelve el último número encontrado, o 0 si el lote está vacío.
    """
    if df_bd_maestra.empty or "ID" not in df_bd_maestra.columns:
        return 0

    # Filtra el DataFrame para obtener solo los registros del lote deseado
    df_lote = df_bd_maestra[df_bd_maestra['ID'].str.startswith(f"{letra_lote}-", na=False)].copy()

    if df_lote.empty:
        return 0 # No hay registros en este lote, empezamos desde cero

    # Extrae la parte numérica, convierte a número y encuentra el máximo
    df_lote['numero_id'] = pd.to_numeric(df_lote['ID'].str.split('-').str[1], errors='coerce')

    ultimo_numero = df_lote['numero_id'].max()

    return 0 if pd.isna(ultimo_numero) else int(ultimo_numero)

def extraer_cp_y_direccion_de_texto(texto_direccion_completa):
    """
    Separa la dirección y el Código Postal de una cadena de texto.
    El CP se espera al final y debe ser de 5 dígitos.
    Devuelve: (direccion_sin_cp, cp_extraido)
    """
    if texto_direccion_completa is None or pd.isna(texto_direccion_completa) or not isinstance(texto_direccion_completa, str) or not texto_direccion_completa.strip():
        return "", "SIN_CP_EN_FUENTE"
    
    cp_match = re.search(r'\b(\d{5})\s*$', texto_direccion_completa.strip()) 
    
    if cp_match:
        cp_extraido = cp_match.group(1)
        direccion_sin_cp = texto_direccion_completa[:cp_match.start()].strip()
        direccion_sin_cp = re.sub(r'[\s,-]+$', '', direccion_sin_cp) # Limpiar caracteres al final
        return direccion_sin_cp, cp_extraido
    else:
        return texto_direccion_completa.strip(), "SIN_CP_VALIDO"
    

def actualizar_o_agregar_registro_bd_unificada(df_bd, registro_data, mode_config):
    id_column_name = mode_config["col_expediente"]
    defined_columns = mode_config["db_master_columns"]
    defined_types = mode_config["db_master_types"]
    id_valor_buscado = str(registro_data.get(id_column_name, "")).strip()

    if not id_valor_buscado:
        print(f"    - (!) Error: Se intentó actualizar/agregar un registro sin ID ('{id_column_name}'). Datos: {registro_data}")
        return df_bd

    if id_column_name not in df_bd.columns:
        print(f"    - (!) Error Crítico: Columna ID '{id_column_name}' no existe en el DataFrame de BD. No se puede actualizar.")
        df_bd[id_column_name] = "" 
        df_bd[id_column_name] = df_bd[id_column_name].astype(str)
    else:
        df_bd[id_column_name] = df_bd[id_column_name].astype(str).str.strip()

    indices_existentes = df_bd.index[df_bd[id_column_name] == id_valor_buscado].tolist()

    if indices_existentes:
        idx_actualizar = indices_existentes[-1]
        for columna, valor_nuevo in registro_data.items():
            if columna in df_bd.columns:
                tipo_esperado = defined_types.get(columna)
                try:
                    valor_convertido = None
                    if pd.isna(valor_nuevo):
                        valor_convertido = pd.NA if tipo_esperado == 'Int64' else ""
                    elif tipo_esperado == 'Int64':
                        num_val = pd.to_numeric(valor_nuevo, errors='coerce')
                        valor_convertido = pd.NA if pd.isna(num_val) else int(num_val)
                    elif tipo_esperado == 'str':
                        valor_convertido = str(valor_nuevo).strip() if valor_nuevo is not None else ""
                    else: 
                        if tipo_esperado in ['float64', 'float32', 'int32'] and isinstance(valor_nuevo, str):
                             valor_convertido = pd.to_numeric(valor_nuevo.replace('',str(pd.NA)), errors='coerce')
                             if pd.notna(valor_convertido): valor_convertido = valor_convertido.astype(tipo_esperado) #type:ignore
                             else: valor_convertido = pd.NA 
                        else:
                            valor_convertido = pd.Series([valor_nuevo]).astype(tipo_esperado).iloc[0]
                    df_bd.loc[idx_actualizar, columna] = valor_convertido
                except Exception as e_conv_update:
                    print(f"      - (!) Advertencia (Actualizar BD): Error al convertir valor '{valor_nuevo}' para col '{columna}' (ID: {id_valor_buscado}). Tipo esperado: {tipo_esperado}. Usando string. Error: {e_conv_update}")
                    df_bd.loc[idx_actualizar, columna] = str(valor_nuevo) if valor_nuevo is not None else ""
    else:
        nueva_fila_dict = {}
        for col_maestra in defined_columns:
            tipo_col_maestra = defined_types.get(col_maestra)
            if col_maestra in registro_data:
                valor_entrante = registro_data[col_maestra]
                try:
                    if pd.isna(valor_entrante):
                        nueva_fila_dict[col_maestra] = pd.NA if tipo_col_maestra == 'Int64' else ""
                    elif tipo_col_maestra == 'Int64':
                        num_val_nuevo = pd.to_numeric(valor_entrante, errors='coerce')
                        nueva_fila_dict[col_maestra] = pd.NA if pd.isna(num_val_nuevo) else int(num_val_nuevo)
                    elif tipo_col_maestra == 'str':
                        nueva_fila_dict[col_maestra] = str(valor_entrante).strip() if valor_entrante is not None else ""
                    else: 
                        if tipo_col_maestra in ['float64', 'float32', 'int32'] and isinstance(valor_entrante, str):
                             converted_val_new = pd.to_numeric(valor_entrante.replace('',str(pd.NA)), errors='coerce')
                             nueva_fila_dict[col_maestra] = converted_val_new.astype(tipo_col_maestra) if pd.notna(converted_val_new) else pd.NA #type:ignore
                        else:
                             nueva_fila_dict[col_maestra] = pd.Series([valor_entrante]).astype(tipo_col_maestra).iloc[0]
                except Exception as e_conv_new_reg:
                    print(f"      - (!) Advertencia (Nuevo Registro BD): Error al convertir valor '{valor_entrante}' para '{col_maestra}'. Tipo esperado: {tipo_col_maestra}. Usando fallback. Error: {e_conv_new_reg}")
                    if tipo_col_maestra == 'Int64': nueva_fila_dict[col_maestra] = pd.NA
                    else: nueva_fila_dict[col_maestra] = str(valor_entrante) if valor_entrante is not None else ""
            else: 
                nueva_fila_dict[col_maestra] = pd.NA if tipo_col_maestra == 'Int64' else ""
        df_nueva_fila = pd.DataFrame([nueva_fila_dict], columns=defined_columns)
        try:
            df_nueva_fila = df_nueva_fila.astype(defined_types)
        except Exception as e_astype_concat:
            print(f"      - (!) Advertencia: Falla en astype antes de concat para nuevo registro ID {id_valor_buscado}. Error: {e_astype_concat}.")
        df_bd = pd.concat([df_bd, df_nueva_fila], ignore_index=True)
        try:
            df_bd = df_bd.astype(defined_types)
        except Exception as e_final_astype:
             print(f"      - (!) Advertencia: Falla en astype final después de concat para ID {id_valor_buscado}. Error: {e_final_astype}.")
    return df_bd

def limpiar_texto(texto_original):
    """
    Limpia un texto para ser usado en nombres de archivo o identificadores.
    Quita acentos, reemplaza caracteres no alfanuméricos (excepto guiones y espacios)
    por guion bajo, y luego reemplaza espacios por guion bajo.
    """
    if texto_original is None or pd.isna(texto_original):
        return "VACIO"
    
    texto_str = str(texto_original)
    
    # Quitar acentos
    try:
        nfkd_form = unicodedata.normalize('NFKD', texto_str)
        texto_sin_acentos = "".join([c for c in nfkd_form if not unicodedata.combining(c)])
    except NameError: # Fallback si unicodedata no está disponible
        texto_sin_acentos = texto_str
        print("    - (!) Advertencia: Módulo 'unicodedata' no disponible. La limpieza de acentos puede no ser completa.")

    # Reemplazar caracteres no alfanuméricos (excepto guiones) por guion bajo
    texto_limpio_paso1 = re.sub(r'[^\w\s\-]', '_', texto_sin_acentos, flags=re.UNICODE)
    # Reemplazar múltiples espacios o guiones bajos por uno solo y quitar los de los extremos
    texto_limpio_paso2 = re.sub(r'[\s_]+', '_', texto_limpio_paso1).strip('_')
    
    return texto_limpio_paso2 if texto_limpio_paso2 else "VACIO"

def convertir_a_formato_con_barras(oficio_str_normalizado):
    """
    Convierte un oficio en formato normalizado (ej. DIDCFMT12345)
    al formato con barras (ej. DI/DCF/MT/12345).
    Si el formato de entrada no es el esperado, devuelve el string original.
    """
    if isinstance(oficio_str_normalizado, str) and oficio_str_normalizado.startswith("DIDCFMT") and oficio_str_normalizado[7:].isdigit():
        return f"DI/DCF/MT/{oficio_str_normalizado[7:]}"
    # Si ya tiene barras o es un formato no reconocido, devolverlo tal cual.
    # Esto es para evitar doble conversión si accidentalmente se le pasa un formato con barras.
    if isinstance(oficio_str_normalizado, str) and re.match(r'DI\s*/\s*DCF\s*/\s*MT\s*/\s*\d+', oficio_str_normalizado):
         return oficio_str_normalizado # Ya está en formato con barras
    return oficio_str_normalizado # Devolver original si no coincide con DIDCFMT...

def cargar_flotillas_xlsx(mode_config):
    """
    Carga la lista de Oficios prioritarios desde la hoja 'Flotillas'
    en el archivo BASE_DE_DATOS.xlsx.
    Devuelve un conjunto (set) de oficios prioritarios.
    """
    filepath = Path(mode_config["data_file_path"]) # Usa el mismo archivo de datos
    sheet_name = "Flotillas"
    flotillas_oficios_set = set()

    if not filepath.exists():
        # Este caso ya se maneja en la carga principal, pero es buena práctica.
        return flotillas_oficios_set

    try:
        # Verificar si la hoja existe sin causar un error fatal si no está
        excel_file = pd.ExcelFile(filepath)
        if sheet_name not in excel_file.sheet_names:
            print(f"    (*) Info para modo '{mode_config['mode_name']}': Hoja '{sheet_name}' no encontrada. Se continuará sin lista de flotillas.")
            return flotillas_oficios_set
        
        print(f"    - Cargando lista de Flotillas desde: {filepath.name} (Hoja: {sheet_name})")
        df_flotillas = pd.read_excel(filepath, sheet_name=sheet_name, dtype=str, keep_default_na=False, na_filter=False)
        df_flotillas.columns = [str(col).strip().upper() for col in df_flotillas.columns]

        oficio_col_name = "OFICIO"
        if oficio_col_name not in df_flotillas.columns:
            print(f"      - (!) Error: Columna '{oficio_col_name}' no encontrada en la hoja '{sheet_name}'. No se cargarán oficios de flotillas.")
            return flotillas_oficios_set

        # Limpiar y obtener los oficios no vacíos
        valid_oficios = df_flotillas[df_flotillas[oficio_col_name].astype(str).str.strip() != ''][oficio_col_name]
        flotillas_oficios_set.update(valid_oficios.astype(str).str.strip())
        
        print(f"      - {len(flotillas_oficios_set)} oficios de Flotillas cargados para priorización.")
        return flotillas_oficios_set
    except Exception as e:
        print(f"    - (!) Error crítico leyendo Flotillas desde '{filepath.name}': {e}")
        return flotillas_oficios_set

def cargar_pm_xlsx(mode_config):
    """
    Carga la lista de Expedientes PM desde el archivo PM.xlsx especificado en mode_config.
    Devuelve un conjunto (set) de expedientes PM.
    """
    filepath = Path(mode_config["pm_file_path"])
    pm_expedientes_set = set()

    if not filepath.exists():
        print(f"    (*) Advertencia para modo '{mode_config['mode_name']}': Archivo PM '{filepath.name}' no encontrado. Se continuará sin lista PM.")
        return pm_expedientes_set
    try:
        print(f"    - Cargando lista PM desde: {filepath.name}")
        # Asumimos que PM.xlsx tiene columnas "EXPEDIENTE" y "TIPO PM"
        # y que queremos los "EXPEDIENTE" donde "TIPO PM" no esté vacío.
        df_pm = pd.read_excel(filepath, dtype=str, keep_default_na=False, na_filter=False)
        df_pm.columns = [str(col).strip().upper() for col in df_pm.columns] # Normalizar encabezados

        exp_col_name = "EXPEDIENTE" # Ajustar si es diferente en tu PM.xlsx
        tipo_pm_col_name = "TIPO PM"   # Ajustar si es diferente en tu PM.xlsx

        if exp_col_name not in df_pm.columns:
            print(f"      - (!) Error: Columna '{exp_col_name}' no encontrada en '{filepath.name}'. No se cargarán expedientes PM.")
            return pm_expedientes_set
        if tipo_pm_col_name not in df_pm.columns:
            # Decidir si esto es un error o si simplemente se cargan todos los expedientes si "TIPO PM" no existe.
            # Por ahora, lo trataremos como que se necesitan ambas columnas.
            print(f"      - (!) Error: Columna '{tipo_pm_col_name}' no encontrada en '{filepath.name}'. No se cargarán expedientes PM.")
            return pm_expedientes_set

        # Filtrar donde TIPO PM no está vacío y obtener los expedientes
        pm_entries_df = df_pm[df_pm[tipo_pm_col_name].astype(str).str.strip() != '']
        pm_expedientes_set.update(pm_entries_df[exp_col_name].astype(str).str.strip())
        
        print(f"      - {len(pm_expedientes_set)} expedientes PM cargados desde '{filepath.name}'.")
        return pm_expedientes_set
    except Exception as e:
        print(f"    - (!) Error crítico leyendo PM desde '{filepath.name}': {e}")
        return pm_expedientes_set
    
def guardar_bd_maestra_unificada(df_a_guardar, mode_config, is_test_mode=False, processed_ids_in_batch=None):
    """
    [VERSIÓN FINAL CON ORDENAMIENTO POR ID] Guarda el DataFrame en el archivo Excel local.
    Incluye un ordenamiento final y robusto por la columna ID para asegurar la consistencia del lote.
    También preserva columnas extra que puedan existir en el Excel (ej. 'MOVIMIENTO').
    """
    global excel_lock
    if is_test_mode:
        print(f"    - [MODO PRUEBA] Guardado de BD para '{mode_config['mode_name']}' OMITIDO.")
        return True

    ruta_excel_bd = Path(mode_config["master_db_file_path"])
    sheet_name_bd = "BD_Maestra"
    defined_columns = mode_config["db_master_columns"]
    defined_types = mode_config["db_master_types"]

    with excel_lock:
        print(f"    - Preparando para guardar BD Maestra para '{mode_config['mode_name']}' ({len(df_a_guardar)} regs) en: '{ruta_excel_bd}'")
        try:
            df_listo_guardar = df_a_guardar.copy()
            
            # --- 1. PREPARACIÓN DE COLUMNAS Y TIPOS ---
            # Capturar todas las columnas presentes en el DataFrame que se va a guardar
            all_columns_in_df = list(df_listo_guardar.columns)
            
            # Crear una lista de columnas final para el guardado (definidas + extras)
            final_columns_for_save = list(defined_columns)
            for col in all_columns_in_df:
                if col not in final_columns_for_save:
                    final_columns_for_save.append(col)

            # Asegurar que todas las columnas definidas existan
            for col_name in defined_columns:
                if col_name not in df_listo_guardar.columns:
                    tipo_esperado_col = defined_types.get(col_name, 'str')
                    df_listo_guardar[col_name] = pd.NA if tipo_esperado_col == 'Int64' else ""
            
            # Reordenar y asegurar tipos
            df_listo_guardar = df_listo_guardar.reindex(columns=final_columns_for_save)
            for col_name, expected_dtype in defined_types.items():
                if col_name in df_listo_guardar.columns:
                    try:
                        if df_listo_guardar[col_name].dtype.name != expected_dtype or expected_dtype == 'str':
                            if expected_dtype == 'Int64':
                                df_listo_guardar[col_name] = pd.to_numeric(df_listo_guardar[col_name], errors='coerce').astype('Int64')
                            elif expected_dtype in ['float64', 'float32']:
                                 df_listo_guardar[col_name] = pd.to_numeric(df_listo_guardar[col_name], errors='coerce').astype(expected_dtype)
                            else:
                                df_listo_guardar[col_name] = df_listo_guardar[col_name].fillna("").astype(str)
                    except Exception as e:
                        print(f"      - (!) Advertencia (guardar): No se pudo convertir '{col_name}' a '{expected_dtype}'. Usando string. Error: {e}")
                        df_listo_guardar[col_name] = df_listo_guardar[col_name].fillna("").astype(str)

            ### INICIO: LÓGICA DE ORDENAMIENTO FINAL POR ID ###
            print("    - Ordenando la base de datos completa por ID de lote antes de guardar...")
            # Solo intentar ordenar si la columna ID existe y tiene al menos un valor no nulo
            if 'ID' in df_listo_guardar.columns and not df_listo_guardar['ID'].isnull().all():
                # Crear columnas temporales para un ordenamiento robusto
                # Se manejan los casos donde un ID pueda ser inválido o estar vacío
                id_parts = df_listo_guardar['ID'].astype(str).str.split('-', n=1, expand=True)
                df_listo_guardar['__letra_lote__'] = id_parts[0].str.upper()
                df_listo_guardar['__numero_lote__'] = pd.to_numeric(id_parts[1], errors='coerce')
                df_listo_guardar['__len_letra__'] = df_listo_guardar['__letra_lote__'].str.len()

                # Ordenar por longitud de la letra, luego la letra, y finalmente el número
                # Esto asegura que 'Z' venga antes de 'AA'
                df_listo_guardar.sort_values(
                    by=['__len_letra__', '__letra_lote__', '__numero_lote__'],
                    ascending=[True, True, True],
                    inplace=True,
                    na_position='last' # Poner filas sin ID válido al final
                )

                # Eliminar las columnas temporales que usamos para ordenar
                df_listo_guardar.drop(columns=['__letra_lote__', '__numero_lote__', '__len_letra__'], inplace=True)
                print("    - Ordenamiento final por ID aplicado.")
            else:
                print("    - No se aplicó ordenamiento (columna ID no presente o vacía).")
            ### FIN: LÓGICA DE ORDENAMIENTO FINAL POR ID ###

            # --- 3. Preparación del Archivo Excel ---
            book = load_workbook(ruta_excel_bd) if ruta_excel_bd.exists() else Workbook()
            ws = book[sheet_name_bd] if sheet_name_bd in book.sheetnames else book.create_sheet(title=sheet_name_bd)
            ws.delete_rows(1, ws.max_row + 1)
            for r in dataframe_to_rows(df_listo_guardar, index=False, header=True):
                cleaned_row = [None if pd.isna(value) else value for value in r]
                ws.append(cleaned_row)
            if "Sheet" in book.sheetnames and sheet_name_bd != "Sheet":
                del book["Sheet"]

            # --- 4. Bucle de Guardado, Reintento y Subida (sin cambios) ---
            max_retries = 3
            delay = 5
            while True:
                for intento in range(max_retries):
                    try:
                        book.save(ruta_excel_bd)
                        print(f"\n    -> BD Maestra local guardada exitosamente.")
                        
                        if not is_test_mode:
                            subir_archivo_al_servidor(ruta_excel_bd, mode_config)
                        
                        return True
                    except PermissionError:
                        if (intento + 1) < max_retries:
                            print(f"\r    - (!) Permiso denegado. Reintentando en {delay}s... ({intento + 1}/{max_retries})", end="")
                            sys.stdout.flush()
                            time.sleep(delay)
                        else:
                            print(f"\r    - (!) Permiso denegado. Todos los {max_retries} reintentos automáticos fallaron.")
                            break
                
                print("\n" + "="*80)
                print(f"    ADVERTENCIA: No se pudo guardar el archivo '{ruta_excel_bd.name}'.")
                print("    Puede estar abierto en Excel, en uso por Power BI, o bloqueado.")
                print("="*80)
                
                respuesta = input("    Presione [Enter] para reintentar o [C] para cancelar: ").strip().lower()
                if respuesta == 'c':
                    print("    -> Operación de guardado cancelada por el usuario.")
                    return False
                print("    -> Reintentando guardado...")

        except Exception as e:
            print(f"    - (!) ERROR CRÍTICO INESPERADO al guardar BD Maestra: {e}")
            traceback.print_exc()
            return False
        
def cargar_config_columnas_xlsx(mode_config):
    filepath = Path(mode_config["config_cols_file_path"])
    print(f"    - Cargando configuración de columnas desde: {filepath.name}")
    df_config = None
    error_msg = "No se pudo leer el archivo de configuración de columnas."
    map_csv_to_maestra = {} # Inicializar por si acaso

    try:
        df_config = pd.read_excel(filepath, dtype=str, keep_default_na=False, na_filter=False)
        df_config.columns = df_config.columns.str.strip()
        
        required_cols_in_config_file = ["NombreEncabezado", "TipoDato"]
        
        if all(col in df_config.columns for col in required_cols_in_config_file):
            df_config["NombreEncabezado"] = df_config["NombreEncabezado"].astype(str).str.strip()
            df_config["TipoDato"] = df_config["TipoDato"].astype(str).str.strip().str.lower()
            
            if "NombreColumnaMaestra" not in df_config.columns:
                print(f"      - ADVERTENCIA: La columna 'NombreColumnaMaestra' no se encontró en '{filepath.name}'. La funcionalidad de actualización de datos en la BD Maestra estará limitada.")
                df_config["NombreColumnaMaestra"] = "" 
            else:
                df_config["NombreColumnaMaestra"] = df_config["NombreColumnaMaestra"].astype(str).str.strip()

            df_config_valid = df_config[(df_config["NombreEncabezado"] != "") & (df_config["TipoDato"] != "")].copy()
            
            if df_config_valid["NombreEncabezado"].duplicated().any():
                print("      - Advertencia: Nombres de encabezado duplicados en config_columnas.xlsx. Se usarán los primeros encontrados.")
                df_config_valid.drop_duplicates(subset="NombreEncabezado", keep='first', inplace=True)
            
            nombres = df_config_valid["NombreEncabezado"].tolist()
            tipos = df_config_valid.set_index('NombreEncabezado')['TipoDato'].to_dict()
            
            # Crear el mapeo: NombreEncabezado (del CSV) -> NombreColumnaMaestra (de la BD Maestra)
            # Solo incluye aquellos donde NombreColumnaMaestra no está vacío.
            map_csv_to_maestra = df_config_valid[df_config_valid["NombreColumnaMaestra"] != ""] \
                                     .set_index('NombreEncabezado')['NombreColumnaMaestra'].to_dict()
            
            print(f"      - Configuración de {len(nombres)} columnas cargada. {len(map_csv_to_maestra)} mapeos a BD Maestra definidos.")
            return nombres, tipos, map_csv_to_maestra # Devolver el nuevo mapeo
        else:
            missing = [col for col in required_cols_in_config_file if col not in df_config.columns]
            error_msg = f"Columnas requeridas {missing} no encontradas en '{filepath.name}'. Columnas detectadas: {df_config.columns.tolist()}"
            
    except FileNotFoundError:
        print(f"    - (!) Error: Archivo de configuración de columnas '{filepath}' no encontrado.")
        return None, None, None 
    except Exception as e:
        error_msg = f"Error leyendo el archivo de configuración de columnas '{filepath.name}': {e}"
    
    print(f"    - (!) Error: {error_msg}")
    return None, None, None

def cargar_datos_principales_xlsx(mode_config, nombres_columnas_config):
    filepath = Path(mode_config["data_file_path"])
    print(f"    - Leyendo datos principales desde: {filepath.name}")
    if not nombres_columnas_config:
        print("    - (!) Error: No se proporcionaron nombres de columna desde la configuración. No se puede leer el archivo de datos principal.")
        return None
    try:
        df_datos = pd.read_excel(
            filepath,
            header=None, 
            names=nombres_columnas_config, 
            skiprows=1, 
            dtype=str,
            keep_default_na=False,
            na_filter=False
        )
        df_datos = df_datos.fillna("")
        print(f"      - Datos principales leídos desde '{filepath.name}'. {len(df_datos)} filas cargadas.")
        return df_datos
    except FileNotFoundError:
        print(f"    - (!) Error: Archivo de datos principal '{filepath}' no encontrado.")
        return None
    except ValueError as ve:
        print(f"    - (!) Error de Valor al leer '{filepath.name}': {ve}. Verifique que 'config_columnas.xlsx' defina el número correcto de columnas para los datos.")
        return None
    except Exception as e:
        print(f"    - (!) Error crítico leyendo datos principales desde '{filepath.name}': {e}")
        return None
# --- FIN FUNCIONES PASO 2 ---

def actualizar_bd_maestra_interactivo(config_modo):
    """
    [VERSIÓN MEJORADA Y COMPLETA] Permite al usuario actualizar registros específicos en la BD Maestra
    buscando la información más reciente en el archivo de datos de origen. Usa la configuración
    dinámica del modo y mantiene características clave como backups.
    """
    print("\n" + "="*25 + f" ACTUALIZACIÓN BD MAESTRA: {config_modo['mode_name']} " + "="*25)

    # 1. Obtener rutas y configuraciones del modo
    ruta_maestra = Path(config_modo["master_db_file_path"])
    ruta_origen = Path(config_modo["data_file_path"])
    llave_principal_maestra = config_modo["col_expediente"]
    llave_principal_origen = config_modo.get("id_col_csv", llave_principal_maestra)

    # 2. Cargar datos de forma robusta
    try:
        print("\nCargando archivos... por favor espera.")
        df_maestra = cargar_bd_maestra_unificada(config_modo)
        if df_maestra is None: return

        nombres_origen, _, mapa_origen_a_maestra = cargar_config_columnas_xlsx(config_modo)
        if not nombres_origen:
            print("(!) Error: No se pudo cargar la configuración de columnas del origen. Abortando.")
            return

        df_origen = cargar_datos_principales_xlsx(config_modo, nombres_origen)
        if df_origen is None or df_origen.empty:
            print(f"(!) El archivo de datos de origen '{ruta_origen.name}' está vacío o no se pudo cargar.")
            return

        df_maestra[llave_principal_maestra] = df_maestra[llave_principal_maestra].astype(str).str.strip()
        df_origen[llave_principal_origen] = df_origen[llave_principal_origen].astype(str).str.strip()
    except Exception as e:
        print(f"(!) Ocurrió un error inesperado al leer los archivos: {e}")
        traceback.print_exc()
        return

    # 3. Solicitar IDs y procesar
    print(f"\nIntroduce los '{llave_principal_maestra}' que deseas actualizar desde '{ruta_origen.name}'.")
    ids_str = input("Puedes pegar una lista separada por comas (ej: 11,22,33): ")
    ids_para_actualizar = [item.strip() for item in ids_str.split(',') if item.strip()]
    if not ids_para_actualizar:
        print("No se ingresaron IDs."); return

    actualizados_count = 0
    no_encontrados_origen = []

    print("\n--- Iniciando Proceso de Actualización ---")
    df_maestra_para_actualizar = df_maestra.copy()

    for id_valor in ids_para_actualizar:
        # Buscar el registro en el archivo de origen
        datos_origen_encontrados = df_origen[df_origen[llave_principal_origen] == id_valor]
        if datos_origen_encontrados.empty:
            print(f"  - (!) Advertencia: '{id_valor}' no fue encontrado en el archivo de Origen. Se omitirá.")
            no_encontrados_origen.append(id_valor)
            continue
        
        print(f"  - Encontrado: {id_valor}. Preparando actualización...")
        fila_origen = datos_origen_encontrados.iloc[0]
        
        # Preparar diccionario para la actualización.
        registro_para_actualizar = {llave_principal_maestra: id_valor}

        # Mapear datos desde el origen a la maestra usando el mapa de config_columnas
        for col_origen, col_maestra in mapa_origen_a_maestra.items():
            if col_origen in fila_origen:
                registro_para_actualizar[col_maestra] = fila_origen[col_origen]
        
        # Lógica específica para MULTAS (cálculos)
        if config_modo['mode_type'] == "MULTAS":
            num_registros = len(datos_origen_encontrados)
            num_hojas = math.ceil(num_registros / 15) # Asumiendo 15 registros por hoja
            registro_para_actualizar['REGISTROS EN BD GENERACION'] = num_registros
            registro_para_actualizar['HOJAS POR DOCUMENTO'] = num_hojas
            print(f"    - (Multas) REGISTROS: {num_registros}, HOJAS: {num_hojas}")

        # Lógica común de estado
        registro_para_actualizar['ESTADO'] = "Actualizado Manualmente"
        registro_para_actualizar['FECHA IMPRESION'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Aplicar la actualización al DataFrame
        df_maestra_para_actualizar = actualizar_o_agregar_registro_bd_unificada(
            df_maestra_para_actualizar, registro_para_actualizar, config_modo
        )
        actualizados_count += 1

    # 4. Guardar los cambios si hubo actualizaciones
    if actualizados_count > 0:
        print(f"\nSe actualizarán {actualizados_count} registro(s).")
        try:
            # Crear un backup antes de sobrescribir
            backup_path = ruta_maestra.with_suffix(f'.{datetime.now().strftime("%Y%m%d_%H%M%S")}.bak')
            print(f"Creando backup de la BD Maestra en: '{backup_path.name}'")
            shutil.copy(ruta_maestra, backup_path)

            print(f"Guardando cambios en '{ruta_maestra.name}'...")
            # Llamamos a la función de guardado unificada
            if guardar_bd_maestra_unificada(df_maestra_para_actualizar, config_modo):
                print("¡Actualización guardada con éxito!")
            else:
                print("(!) ERROR CRÍTICO AL GUARDAR LA BD MAESTRA.")
                print("Tus cambios NO han sido guardados. Revisa el archivo de backup.")
        except Exception as e:
            print(f"\n(!) ERROR CRÍTICO DURANTE EL PROCESO DE GUARDADO: {e}")
            traceback.print_exc()
    else:
        print("\nNo se actualizó ningún registro.")
    
    input("Presiona Enter para volver al menú...")

def detectar_modos_disponibles():
    modos_encontrados = {}
    print(f"\nBuscando modos de operación en: {SCRIPT_BASE_PATH}")
    for item in SCRIPT_BASE_PATH.iterdir():
        if item.is_dir():
            mode_name_candidate = item.name
            mode_type_detected = None
            current_mode_config = {}

            if KEYWORD_MULTAS.lower() in mode_name_candidate.lower():
                mode_type_detected = "MULTAS"
            elif KEYWORD_PREDIAL.lower() in mode_name_candidate.lower():
                mode_type_detected = "PREDIAL"

            if mode_type_detected:
                print(f"  - Candidato a modo: '{mode_name_candidate}' (Tipo: {mode_type_detected})")
                base_path_modo = item
                data_file = base_path_modo / "BASE_DE_DATOS.xlsx"
                config_cols_file = base_path_modo / "config_columnas.xlsx"
                pm_file = base_path_modo / "PM.xlsx"
                template_file = base_path_modo / "PLANTILLA.docx"
                
                # Construir nombre de la BD Maestra
                # Para Multas, es fijo. Para Predial, depende del nombre de la carpeta.
                if mode_type_detected == "MULTAS":
                    master_db_filename = "base_datos_maestra_multas.xlsx"
                else: # PREDIAL
                    # Limpiar nombre de carpeta para usarlo en el nombre de la BD maestra
                    # ej. "PREDIAL_FERNANDO" -> "base_datos_maestra_predial_fernando.xlsx"
                    clean_mode_name = re.sub(r'[^\w_]', '', mode_name_candidate.lower())
                    master_db_filename = f"base_datos_maestra_{clean_mode_name}.xlsx"
                
                master_db_file = base_path_modo / master_db_filename

                if not data_file.exists():
                    print(f"    (!) OMITIENDO MODO '{mode_name_candidate}': Falta '{data_file.name}'.")
                    continue
                if not config_cols_file.exists():
                    print(f"    (!) OMITIENDO MODO '{mode_name_candidate}': Falta '{config_cols_file.name}'.")
                    continue
                if not template_file.exists():
                    print(f"    (!) OMITIENDO MODO '{mode_name_candidate}': Falta '{template_file.name}'.")
                    continue
                if not pm_file.exists(): # PM es opcional pero útil registrarlo
                     print(f"    (*) Advertencia para modo '{mode_name_candidate}': Falta '{pm_file.name}'.")


                current_mode_config["mode_name"] = mode_name_candidate
                current_mode_config["mode_type"] = mode_type_detected
                current_mode_config["base_path"] = base_path_modo
                current_mode_config["data_file_name"] = data_file.name
                current_mode_config["config_cols_file_name"] = config_cols_file.name
                current_mode_config["pm_file_name"] = pm_file.name
                current_mode_config["template_file_name"] = template_file.name
                current_mode_config["master_db_file_name"] = master_db_file.name # Nombre construido
                current_mode_config["output_docs_folder_name"] = "DOCUMENTOS"

                current_mode_config["data_file_path"] = data_file
                current_mode_config["config_cols_file_path"] = config_cols_file
                current_mode_config["pm_file_path"] = pm_file
                current_mode_config["template_file_path"] = template_file
                current_mode_config["master_db_file_path"] = master_db_file
                current_mode_config["output_docs_path"] = base_path_modo / current_mode_config["output_docs_folder_name"]

                if mode_type_detected == "PREDIAL":
                    current_mode_config["db_master_columns"] = COLUMNAS_BD_MAESTRA_PREDIAL
                    current_mode_config["db_master_types"] = TIPOS_BD_MAESTRA_PREDIAL
                    current_mode_config["col_expediente"] = COL_ID_PREDIAL # ID en BD Maestra y en CSV
                    current_mode_config["col_estado_bd_maestra"] = COL_ESTADO_PREDIAL_BD
                    current_mode_config["col_nombre_csv_original"] = COL_NOMBRE_CSV_PREDIAL
                    current_mode_config["col_colonia_csv_original"] = COL_COLONIA_CSV_PREDIAL
                    current_mode_config["col_periodo_csv_predial"] = "PERIODO"
                    current_mode_config["col_base_datos_escaneo_bd_maestra"] = COL_BASE_DATOS_ESCANEO_PREDIAL # Opcional, si la lógica lo necesita explícitamente

                    # --- NUEVAS CONFIGURACIONES PARA EL RECORTE DE OCR PREDIA ---
                    # General
                    current_mode_config["ocr_zoom_factor_predial"] = 2.5 # Puedes probar con 2.0, 2.5 o 3.0

                    # Para la parte SUPERIOR (donde está "EXP 84425002")
                    current_mode_config["ocr_predial_top_start_y_perc"] = 0.05  # Empezar a buscar desde el 5% de la altura (para evitar el borde mismo)
                    current_mode_config["ocr_predial_top_end_y_perc"] = 0.25    # Terminar de buscar en el 25% de la altura (ajusta si EXP está más abajo)
                    current_mode_config["ocr_tesseract_config_predial_top"] = "--psm 4" # Probar con psm 4 (columna única) o 3 (automático)

                    # Para la parte INFERIOR (donde está "EXP. 84425002" en el pie de página)
                    current_mode_config["ocr_predial_bottom_start_y_perc"] = 0.88 # Empezar a buscar en el 88% (cerca del pie)
                    current_mode_config["ocr_predial_bottom_end_y_perc"] = 1.0   # Hasta el final de la página (100%)
                    current_mode_config["ocr_tesseract_config_predial_bottom"] = "--psm 6" # psm 6 suele ser bueno para líneas individuales
                    
                    print(f"    -> Modo '{mode_name_candidate}' configurado y añadido.") #



                elif mode_type_detected == "MULTAS":
                    current_mode_config["db_master_columns"] = COLUMNAS_BD_MAESTRA_MULTAS
                    current_mode_config["db_master_types"] = TIPOS_BD_MAESTRA_MULTAS
                    current_mode_config["col_expediente"] = COL_ID_MULTAS_BD # OFICIO en BD Maestra
                    current_mode_config["id_col_csv"] = COL_ID_MULTAS_CSV # OFICIO en BASE_DE_DATOS.xlsx
                    current_mode_config["col_estado_bd_maestra"] = COL_ESTADO_MULTAS_BD
                    current_mode_config["col_nombre_base_csv"] = COL_NOMBRE_CSV_MULTAS_BASE # Para construir nombre
                    current_mode_config["col_apaterno_csv"] = COL_APATERNO_CSV_MULTAS_BASE
                    current_mode_config["col_amaterno_csv"] = COL_AMATERNO_CSV_MULTAS_BASE
                    current_mode_config["col_cp_csv"] = COL_CP_CSV_MULTAS
                    current_mode_config["col_direccion_completa_csv"] = "DIRECCION" 
                    current_mode_config["col_conteo_registros_generacion_maestra"] = COL_CONTEO_REGISTROS_MULTAS # Nombre en BD Maestra




                # Añadir constantes de estado al config del modo para fácil acceso
                for estado_const_name, estado_const_value in globals().items():
                    if estado_const_name.startswith("ESTADO_") or estado_const_name.startswith("MODO_GENERACION_"):
                        current_mode_config[estado_const_name] = estado_const_value
                
                modos_encontrados[mode_name_candidate] = current_mode_config
                print(f"    -> Modo '{mode_name_candidate}' configurado y añadido.")
    
    if not modos_encontrados:
        print("  (!) No se encontraron directorios de modo válidos.")
    return modos_encontrados


def main_generador_loop():
    print("===================================================================")
    print("--- Sistema Unificado de Generación de Documentos (Maestro) ---")
    print(f"    Directorio Base: {SCRIPT_BASE_PATH}")
    print("===================================================================")

    modos_disponibles = detectar_modos_disponibles()
    if not modos_disponibles:
        input("No hay modos de operación válidos. Presione Enter para salir.")
        return

    while True:
        print("\n" + "=" * 30 + " MENÚ PRINCIPAL " + "=" * 30)
        print("Modos de Operación Disponibles:")
        modos_nombres_listados = list(modos_disponibles.keys())
        for i, nombre_modo in enumerate(modos_nombres_listados):
            print(f"  {i+1}. Procesar Modo: {nombre_modo}")
        print("-" * 78)
        # --- NUEVA OPCIÓN DE MENÚ ---
        print("  A. Actualizar TODOS los archivos en el Servidor (Forzar Sincronización)")
        print("  S. SALIR del programa")
        print("=" * 78)

        opcion_modo_idx_str = input("Seleccione una opción: ").strip().upper()

        if opcion_modo_idx_str == 'S':
            print("Saliendo del sistema de generación...")
            break
        
        # --- NUEVA LÓGICA PARA MANEJAR LA OPCIÓN 'A' ---
        elif opcion_modo_idx_str == 'A':
            sincronizar_todo_con_servidor()
            input("\n    Sincronización finalizada. Presione Enter para continuar...")
            continue

        try:
            idx_seleccionado = int(opcion_modo_idx_str) - 1
            if 0 <= idx_seleccionado < len(modos_nombres_listados):
                nombre_modo_seleccionado = modos_nombres_listados[idx_seleccionado]
                config_modo_actual = modos_disponibles[nombre_modo_seleccionado]
                
                output_path_obj = config_modo_actual["output_docs_path"]
                output_path_obj.mkdir(parents=True, exist_ok=True)
                if config_modo_actual["mode_type"] == "PREDIAL":
                    (output_path_obj / "COLONIAS").mkdir(parents=True, exist_ok=True)
                    (output_path_obj / "VACIAS").mkdir(parents=True, exist_ok=True)
                elif config_modo_actual["mode_type"] == "MULTAS":
                    (output_path_obj / "CP").mkdir(parents=True, exist_ok=True)
                    (output_path_obj / "VACIAS").mkdir(parents=True, exist_ok=True)

                sub_menu_acciones_modo(config_modo_actual)
            else:
                print("Opción de modo inválida.")
        except ValueError:
            print("Entrada inválida. Por favor, ingrese un número o una letra de opción válida.")
        
        input("    Presione Enter para continuar...")

    print("\n--- Sistema Unificado de Generación Finalizado. ---")

def process_selected_mode_action(mode_config, modo_accion_solicitada):
    """
    Procesa la acción de generación o escaneo para el modo y tipo de acción seleccionados.
    """
    print(f"\n--- Iniciando Procesamiento para Modo: {mode_config['mode_name']} ({mode_config['mode_type']}) ---")
    print(f"    Acción Solicitada: {modo_accion_solicitada}")

    # --- Pasos 1, 2 y 3: Carga de Datos (Necesarios para todas las acciones) ---
    print("  1. Cargando Base de Datos Maestra del modo...")
    df_bd_maestra_actual = cargar_bd_maestra_unificada(mode_config)
    if df_bd_maestra_actual is None:
        return

    print(f"\n  2. Cargando configuración de columnas desde '{mode_config['config_cols_file_name']}'...")
    nombres_col_csv, tipos_col_csv, _ = cargar_config_columnas_xlsx(mode_config)
    if nombres_col_csv is None:
        return

    print(f"\n  3. Cargando datos principales desde '{mode_config['data_file_name']}'...")
    df_registros_para_logica = cargar_datos_principales_xlsx(mode_config, nombres_col_csv)
    if df_registros_para_logica is None:
        df_registros_para_logica = pd.DataFrame() # Asegurarse de que sea un DataFrame vacío si falla la carga

    # --- INICIO DE LA NUEVA ESTRUCTURA LÓGICA ---
    # EXPLICACIÓN: Se crea una bifurcación para separar los flujos de trabajo de GENERAR y ESCANEAR.
    
    acciones_de_generacion = [
        MODO_GENERACION_COMPLETO, MODO_GENERACION_ULTIMA,
        MODO_GENERACION_RESTO, MODO_GENERACION_ESPECIFICOS
    ]

    df_bd_maestra_actualizada_desde_logica = None
    processed_ids_multas_batch = None

    # --- BLOQUE EXCLUSIVO PARA ACCIONES DE GENERACIÓN ---
    if modo_accion_solicitada in acciones_de_generacion:
        print("\n  -> Iniciando flujo de GENERACIÓN de documentos...")
        
        if df_registros_para_logica.empty:
            print(f"  (*) Información: El archivo de datos '{mode_config['data_file_name']}' está vacío. No hay nada que generar.")
            return
        
        if modo_accion_solicitada in acciones_de_generacion and mode_config["mode_type"] == "MULTAS":
            print("    - Realizando extracción de Código Postal desde la dirección para ordenar...")
            # Obtener el nombre de la columna de dirección desde la configuración del modo
            col_direccion_nombre = mode_config.get("col_direccion_completa_csv")

            if col_direccion_nombre and col_direccion_nombre in df_registros_para_logica.columns:
                # Usar la función existente para extraer el CP y crear la nueva columna 'CP'
                # La función devuelve (direccion, cp), por eso usamos [1] para tomar solo el cp.
                df_registros_para_logica['CP'] = df_registros_para_logica[col_direccion_nombre].apply(
                    lambda x: extraer_cp_y_direccion_de_texto(x)[1]
                )
                print(f"    - Columna 'CP' creada exitosamente a partir de la columna '{col_direccion_nombre}'.")
            else:
                print(f"      - (!) ADVERTENCIA: La columna de dirección '{col_direccion_nombre}' no se encontró.")
                print("      - Se creará una columna 'CP' vacía. El ordenamiento por CP no será efectivo.")
                df_registros_para_logica['CP'] = "SIN_CP"

        # --- LÓGICA DE SELECCIÓN Y ORDENAMIENTO PRECISO (CORREGIDA) ---
        print("\n  4. Identificando y ordenando expedientes a procesar...")
        
        id_col_en_bd = mode_config["col_expediente"]
        id_col_en_datos = mode_config.get("id_col_csv", id_col_en_bd)
        
        # Esta variable contendrá los IDs en el orden correcto.
        ids_ordenados_a_procesar = []

        # Paso A: Obtener la LISTA ORDENADA de IDs.
        if modo_accion_solicitada == MODO_GENERACION_ESPECIFICOS:
            expedientes_str = input(f"    Ingrese los OFICIOS a generar (últimos 6 dígitos), separados por comas: ").strip()
            if not expedientes_str:
                print("    No se ingresaron expedientes. Abortando."); return
            ids_ordenados_a_procesar = [exp.strip().upper() for exp in expedientes_str.split(',') if exp.strip()]
        
        else:
            # Para los demás modos, la fuente de IDs cambia, pero el objetivo es el mismo: obtener una lista ordenada.
            df_fuente_para_ids = pd.DataFrame()
            if modo_accion_solicitada == MODO_GENERACION_RESTO:
                estados_a_buscar = [ESTADO_IMP_ULTIMA, ESTADO_GEN_ULTIMA]
                print(f"    - Modo RESTO. Filtrando BD Maestra por estados: {estados_a_buscar}")
                df_fuente_para_ids = df_bd_maestra_actual[df_bd_maestra_actual['ESTADO'].isin(estados_a_buscar)]
            
            elif modo_accion_solicitada == MODO_GENERACION_ULTIMA:
                estados_a_buscar = [ESTADO_PENDIENTE, ESTADO_ERROR_GENERACION]
                print(f"    - Modo ULTIMA. Filtrando BD Maestra por estados: {estados_a_buscar}")
                df_fuente_para_ids = df_bd_maestra_actual[df_bd_maestra_actual['ESTADO'].isin(estados_a_buscar)]
            
            else: # MODO_GENERACION_COMPLETO
                # --- INICIO DE LA NUEVA LÓGICA INTELIGENTE ---
                print(f"    - Modo COMPLETO (Inteligente). Comparando origen vs. BD Maestra para encontrar expedientes nuevos...")

                # 1. Obtener todos los expedientes únicos que YA EXISTEN en la Base de Datos Maestra.
                #    Se usa un 'set' para una comparación muy rápida.
                expedientes_en_maestra_set = set(df_bd_maestra_actual[id_col_en_bd].astype(str).str.strip())
                print(f"      - Se encontraron {len(expedientes_en_maestra_set)} expedientes únicos en la BD Maestra.")

                # 2. Obtener todos los expedientes únicos del archivo de origen (BASE_DE_DATOS.xlsx).
                expedientes_en_origen_lista = df_registros_para_logica[id_col_en_datos].astype(str).str.strip().unique().tolist()
                
                # 3. Comparar ambas listas para encontrar los expedientes que solo están en el origen (los nuevos).
                #    Se itera sobre la lista del origen para mantener el orden original del archivo.
                expedientes_nuevos_a_procesar = [
                    exp for exp in expedientes_en_origen_lista if exp not in expedientes_en_maestra_set
                ]
                
                # Asignar la lista de expedientes nuevos directamente a la variable que se usará después.
                ids_ordenados_a_procesar = expedientes_nuevos_a_procesar
                print(f"      - {len(expedientes_nuevos_a_procesar)} expedientes nuevos encontrados para generar.")
                # --- FIN DE LA NUEVA LÓGICA INTELIGENTE ---

            if not ids_ordenados_a_procesar: # Se reemplaza la lógica anterior por esta verificación directa
                print(f"    -> No se encontraron expedientes para procesar en el modo '{modo_accion_solicitada}'.")
                return

    # --- INICIO DE LA NUEVA LÓGICA DE ORDENAMIENTO INTELIGENTE (PREDIAL Y MULTAS) ---
        print(f"    -> Se encontraron {len(ids_ordenados_a_procesar)} expedientes/oficios candidatos. Aplicando ordenamiento...")
        # --- INICIO: Bloque para normalizar búsqueda de Oficio ---
        print("    - Normalizando IDs para búsqueda...")
        # Crea una columna temporal que contiene solo la parte numérica del oficio
        df_registros_para_logica['__OFICIO_NUMERICO__'] = df_registros_para_logica[id_col_en_datos].str.extract(r'(\d+)$').fillna('')
        # --- FIN: Bloque de normalización ---

        df_a_ordenar = df_registros_para_logica[df_registros_para_logica['__OFICIO_NUMERICO__'].isin(ids_ordenados_a_procesar)].copy()


        # --- INICIO DE NUEVA VERIFICACIÓN ---
        if df_a_ordenar.empty and modo_accion_solicitada == MODO_GENERACION_ESPECIFICOS:
            print("\n(!) ERROR: Ninguno de los OFICIOS especificados fue encontrado en BASE_DE_DATOS.xlsx.")
            print("    - Revisa que los números de oficio sean correctos y existan en el archivo de datos.")
            return # Detiene la ejecución de esta acción
        # --- FIN DE NUEVA VERIFICACIÓN ---

        if mode_config["mode_type"] == "PREDIAL":
            # Lógica de ordenamiento para PREDIAL (esta parte se mantiene igual)
            col_colonia_orden = mode_config.get("col_colonia_csv_original", "COLONIA")
            df_a_ordenar['__COLONIA_PROC__'] = df_a_ordenar[col_colonia_orden].astype(str).fillna('VACIA').str.strip().upper()
            df_a_ordenar.loc[df_a_ordenar['__COLONIA_PROC__'] == '', '__COLONIA_PROC__'] = 'VACIA'
            
            if 'TOTAL' in df_a_ordenar.columns:
                df_a_ordenar['__MONTO_SORT__'] = pd.to_numeric(df_a_ordenar['TOTAL'].astype(str).str.replace(r'[$,]', '', regex=True), errors='coerce').fillna(0.0)
            else:
                df_a_ordenar['__MONTO_SORT__'] = 0.0

            df_a_ordenar['__CONTEO_COLONIA__'] = df_a_ordenar.groupby('__COLONIA_PROC__')[id_col_en_datos].transform('nunique')
            
            df_para_usar_en_logica_generacion = df_a_ordenar.sort_values(
                by=['__CONTEO_COLONIA__', '__COLONIA_PROC__', '__MONTO_SORT__'],
                ascending=[False, True, False]
            ).drop(columns=['__COLONIA_PROC__', '__MONTO_SORT__', '__CONTEO_COLONIA__'])
        
        elif mode_config["mode_type"] == "MULTAS":
            # Lógica de ordenamiento específica para MULTAS (CP y Monto Total)
            col_cp_orden = "CP"
            col_id_multas = mode_config.get("id_col_csv", "OFICIO")
            col_importe_multas = mode_config.get("col_importe_csv", "IMPORTE")

            # 1. Crear columna numérica de importe para poder sumar (si no existe)
            if col_importe_multas in df_a_ordenar.columns:
                df_a_ordenar['__IMPORTE_NUMERIC__'] = pd.to_numeric(
                    df_a_ordenar[col_importe_multas].astype(str).str.replace(r'[$,]', '', regex=True),
                    errors='coerce'
                ).fillna(0.0)
            else:
                df_a_ordenar['__IMPORTE_NUMERIC__'] = 0.0

            # 2. Calcular el conteo de oficios únicos por Código Postal
            if col_cp_orden in df_a_ordenar.columns:
                df_a_ordenar['__cp_count__'] = df_a_ordenar.groupby(col_cp_orden)[col_id_multas].transform('nunique')
            else:
                df_a_ordenar['__cp_count__'] = 0

            # 3. Calcular el monto total por cada Oficio único
            df_a_ordenar['__monto_total_oficio__'] = df_a_ordenar.groupby(col_id_multas)['__IMPORTE_NUMERIC__'].transform('sum')

            # 4. Eliminar duplicados para tener una lista de oficios únicos ordenados
            df_oficios_unicos_ordenados = df_a_ordenar.drop_duplicates(subset=[col_id_multas]).sort_values(
                by=['__cp_count__', '__monto_total_oficio__'],
                ascending=[False, False]
            )
            
            # 5. Filtrar el DataFrame original para que siga el nuevo orden de oficios
            lista_oficios_ordenada = df_oficios_unicos_ordenados[col_id_multas].tolist()
            df_a_ordenar[col_id_multas] = pd.Categorical(df_a_ordenar[col_id_multas], categories=lista_oficios_ordenada, ordered=True)
            df_para_usar_en_logica_generacion = df_a_ordenar.sort_values(col_id_multas).drop(columns=['__IMPORTE_NUMERIC__', '__cp_count__', '__monto_total_oficio__'])

        else: 
            df_para_usar_en_logica_generacion = df_a_ordenar

        print("    -> Ordenamiento finalizado. Los documentos se generarán en la secuencia correcta.")
        # --- FIN DE LA NUEVA LÓGICA DE ORDENAMIENTO ---

        # Paso D: Preguntar al usuario cuántos generar.
        num_disponibles = len(df_para_usar_en_logica_generacion[id_col_en_datos].unique())
        max_archivos_para_ejecutar = num_disponibles
        if modo_accion_solicitada != MODO_GENERACION_ESPECIFICOS:
            prompt_cantidad = (f"\n    Hay {num_disponibles} expedientes listos en el orden correcto. ¿Cuántos desea generar?\n"
                               f"    (Presione Enter para todos, 0 para ninguno): ")
            respuesta_cantidad = input(prompt_cantidad).strip()
            if respuesta_cantidad:
                try: 
                    max_archivos_para_ejecutar = int(respuesta_cantidad)
                    if max_archivos_para_ejecutar > num_disponibles:
                        print(f"    El número es mayor al disponible, se generarán {num_disponibles}.")
                        max_archivos_para_ejecutar = num_disponibles
                except ValueError: 
                    max_archivos_para_ejecutar = 0
        
        if modo_accion_solicitada != MODO_GENERACION_ESPECIFICOS and max_archivos_para_ejecutar <= 0:
            print("    Generación cancelada por el usuario."); return
        
        # --- LÓGICA DE GENERACIÓN DE ID (Solo para Generación) ---
        letra_lote_actual, contador_lote_actual = obtener_ultimo_id(df_bd_maestra_actual)
        
        if contador_lote_actual >= 100:
            # Si el último lote se llenó (100 o más), empezamos uno nuevo.
            letra_para_siguiente_tanda = obtener_siguiente_letra_lote(letra_lote_actual)
            contador_para_siguiente_tanda = 0
            print(f"    -> Lote '{letra_lote_actual}' está completo. El nuevo lote comenzará con la letra '{letra_para_siguiente_tanda}'.")
        elif letra_lote_actual is None:
            # Si no hay lotes, empezamos con 'A' desde 0.
            letra_para_siguiente_tanda = 'A'
            contador_para_siguiente_tanda = 0
            print("    -> No se encontraron lotes previos. Se iniciará con el lote 'A'.")
        else:
            # Si el último lote no está lleno, continuamos con él.
            letra_para_siguiente_tanda = letra_lote_actual
            contador_para_siguiente_tanda = contador_lote_actual
            print(f"    -> Continuando con el lote '{letra_para_siguiente_tanda}' a partir del número {contador_para_siguiente_tanda}.")



        # --- LLAMADA A LA LÓGICA DE GENERACIÓN ---
        if mode_config["mode_type"] == "PREDIAL":
            pm_set_actual = cargar_pm_xlsx(mode_config)
            df_bd_maestra_actualizada_desde_logica = GeneradorPredial_logica.generar_documentos_predial_core(
                df_datos_para_procesar=df_para_usar_en_logica_generacion, pm_set_actual=pm_set_actual,
                config_predial_actual=mode_config, modo_generacion_solicitado=modo_accion_solicitada,
                max_docs_a_generar=max_archivos_para_ejecutar,
                df_bd_maestra_actualizada=df_bd_maestra_actual,
                nombres_columnas_csv=nombres_col_csv,
                tipos_columnas_csv=tipos_col_csv,
                letra_lote=letra_para_siguiente_tanda,
                contador_inicial_lote=contador_para_siguiente_tanda, # <-- ESTA ES LA LÍNEA CORRECTA
                # La línea repetida ha sido eliminada
                funcion_de_subida=subir_archivo_al_servidor
            )
        elif mode_config["mode_type"] == "MULTAS":
            df_bd_maestra_actualizada_desde_logica, processed_ids_multas_batch = GeneradorMultas_logica.generar_documentos_multas_core(
                df_datos_principales=df_para_usar_en_logica_generacion, df_bd_maestra_actual=df_bd_maestra_actual,
                config_multas_actual=mode_config, modo_generacion_menu_solicitado=modo_accion_solicitada,
                max_archivos_a_generar=max_archivos_para_ejecutar,
                letra_lote=letra_para_siguiente_tanda,
                contador_inicial_lote=contador_para_siguiente_tanda
            )
    # --- BLOQUE EXCLUSIVO PARA ACCIONES DE ESCANEO ---
    elif modo_accion_solicitada in [MODO_ACCION_ESCANEAR_MULTAS, MODO_ACCION_ESCANEAR_PREDIAL]:
        print("\n  -> Iniciando flujo de ESCANEO de documentos...")
        
        # EXPLICACIÓN: Para escanear, usamos los datos tal como se cargaron, sin ordenar ni filtrar por cantidad.
        df_para_usar_en_logica_generacion = df_registros_para_logica.copy()
        
        # --- LLAMADA A LA LÓGICA DE ESCANEO ---
        if mode_config["mode_type"] == "PREDIAL":
            df_bd_maestra_actualizada_desde_logica = GeneradorPredial_logica.run_scan_and_process_predial(
                df_bd_maestra_global=df_bd_maestra_actual,
                df_csv_principal_global=df_para_usar_en_logica_generacion, 
                config_predial_actual=mode_config,
                funcion_de_subida=subir_archivo_al_servidor # <-- PARÁMETRO AÑADIDO
            )
        elif mode_config["mode_type"] == "MULTAS":
            df_bd_maestra_actualizada_desde_logica = GeneradorMultas_logica.run_scan_and_process_multas(
                df_bd_maestra_global=df_bd_maestra_actual,
                df_csv_principal_global=df_para_usar_en_logica_generacion, 
                config_multas_actual=mode_config,
                funcion_de_subida=subir_archivo_al_servidor 
            )
            
    else:
        print(f"  (!) Acción '{modo_accion_solicitada}' no reconocida como de generación o escaneo.")
        return

    # --- 6. Guardado de la Base de Datos (Común a ambos flujos) ---
    if df_bd_maestra_actualizada_desde_logica is not None:
        # Extraer la constante MODO_GENERACION_ESPECIFICOS para la comparación
        MODO_ESPECIFICOS = mode_config.get("MODO_GENERACION_ESPECIFICOS", "ESPECIFICOS")

        # Si el modo es ESPECIFICOS, no se guarda la BD.
        if modo_accion_solicitada == MODO_ESPECIFICOS:
            print("\n[MODO ESPECÍFICOS] Documentos generados. La Base de Datos Maestra no fue modificada, como se solicitó.")
        else:
            # Para todos los demás modos, se guarda y se sube como de costumbre.
            print("\n  6. Guardando Base de Datos Maestra actualizada...")
            guardar_bd_maestra_unificada(
                df_bd_maestra_actualizada_desde_logica,
                mode_config,
                is_test_mode=False,
                processed_ids_in_batch=processed_ids_multas_batch
            )
    print(f"\n--- Procesamiento para Modo: {mode_config['mode_name']} (Acción: {modo_accion_solicitada}) Finalizado ---")

def buscar_pdf_oficial_generado(oficio_norm, config_modo, nombre_contrib, cp):
    """
    Busca activamente un PDF generado oficial basado en la información del oficio.
    Devuelve la ruta (Path object) si lo encuentra, o None si no.
    """
    print(f"        - Buscando PDF oficial para Oficio (Norm): {oficio_norm}...")
    
    # Directorio base donde se guardan los documentos generados
    directorio_documentos = Path(config_modo["output_docs_path"])
    
    # Recrear el nombre de archivo como lo hace el generador
    oficio_con_barras = convertir_a_formato_con_barras(oficio_norm)
    nombre_base_archivo = limpiar_texto(f"{oficio_con_barras}_{nombre_contrib}")
    nombre_archivo_pdf = f"{nombre_base_archivo}.pdf"
    
    # Definir las rutas de búsqueda en orden de prioridad
    rutas_a_buscar = []
    
    # 1. Buscar en la carpeta del Código Postal (si existe)
    if cp and cp not in ["SIN_CP", "SIN_CP_VALIDO"]:
        ruta_en_cp = directorio_documentos / "CP" / f"CP_{cp}" / nombre_archivo_pdf
        rutas_a_buscar.append(ruta_en_cp)

    # 2. Buscar en la carpeta de VACIAS
    ruta_en_vacias = directorio_documentos / "VACIAS" / nombre_archivo_pdf
    rutas_a_buscar.append(ruta_en_vacias)
    
    # 3. Buscar directamente en la carpeta DOCUMENTOS (como respaldo)
    ruta_en_raiz = directorio_documentos / nombre_archivo_pdf
    rutas_a_buscar.append(ruta_en_raiz)

    # Iterar y devolver la primera coincidencia que exista
    for ruta in rutas_a_buscar:
        if ruta.is_file():
            print(f"          -> ¡Éxito! PDF oficial encontrado en: {ruta}")
            return ruta
            
    print(f"        - No se encontró un PDF oficial pre-generado para el oficio {oficio_norm}.")
    return None

def procesar_accion_vista_rapida_predial_ui(mode_config):
    print(f"\n--- [UI] Iniciando VISTA RÁPIDA (Primeras Dos Páginas) para Modo: {mode_config['mode_name']} ---")
    print("    NOTA: Esta opción NO actualiza la Base de Datos Maestra.")

    if mode_config["mode_type"] != "PREDIAL":
        print(f"    Esta función de vista rápida está actualmente configurada solo para modos PREDIAL.")
        input("    Presione Enter para continuar...")
        return

    # 1. Cargar Configuración de Columnas
    print(f"\n  1. Cargando config. de columnas: '{mode_config['config_cols_file_name']}'...")
    nombres_col_csv, tipos_col_csv = cargar_config_columnas_xlsx(mode_config)
    if nombres_col_csv is None:
        print(f"  (!) Error Crítico: No se pudo cargar config de columnas. Abortando vista rápida.")
        input("    Presione Enter para continuar...")
        return

    # 2. Cargar Datos Principales (BASE_DE_DATOS.xlsx)
    print(f"\n  2. Cargando datos principales: '{mode_config['data_file_name']}'...")
    df_datos_originales = cargar_datos_principales_xlsx(mode_config, nombres_col_csv)
    if df_datos_originales is None or df_datos_originales.empty:
        print(f"  (!) Sin datos en '{mode_config['data_file_name']}'. Abortando vista rápida.")
        input("    Presione Enter para continuar...")
        return

    # 3. Ordenar Datos según tus criterios
    print("\n  3. Ordenando registros...")
    col_id_csv_para_orden = mode_config.get("id_col_csv", mode_config["col_expediente"])
    col_colonia_csv_para_orden = mode_config.get("col_colonia_csv_original", "COLONIA")

    df_a_ordenar = df_datos_originales.copy()
    df_a_ordenar['COLONIA_ORD_TEMP'] = df_a_ordenar[col_colonia_csv_para_orden].astype(str).fillna('VACIA').str.strip().str.upper()
    df_a_ordenar.loc[df_a_ordenar['COLONIA_ORD_TEMP'] == '', 'COLONIA_ORD_TEMP'] = 'VACIA'
    df_a_ordenar['CONTEO_COL_TEMP'] = df_a_ordenar.groupby('COLONIA_ORD_TEMP')[col_id_csv_para_orden].transform('nunique')
    df_a_ordenar[col_id_csv_para_orden] = df_a_ordenar[col_id_csv_para_orden].astype(str)
    df_a_ordenar['EXP_NUM_TEMP'] = pd.to_numeric(df_a_ordenar[col_id_csv_para_orden], errors='coerce')

    df_datos_ordenados_vista = df_a_ordenar.sort_values(
        by=['CONTEO_COL_TEMP', 'COLONIA_ORD_TEMP', 'EXP_NUM_TEMP', col_id_csv_para_orden],
        ascending=[False, True, True, True],
        na_position='last'
    ).drop(columns=['COLONIA_ORD_TEMP', 'CONTEO_COL_TEMP', 'EXP_NUM_TEMP'])

    print(f"    Registros ordenados: {len(df_datos_ordenados_vista)}")
    if df_datos_ordenados_vista.empty:
        input("    No hay datos después del ordenamiento. Presione Enter...")
        return

    # 4. Preguntar cuántos generar
    num_max_vista = 0
    try:
        input_num_str = input(f"    ¿Cuántos documentos de la lista ordenada ({len(df_datos_ordenados_vista)}) desea generar para vista rápida? (Enter o 0 para todos): ").strip()
        if not input_num_str or input_num_str == "0":
            num_max_vista = len(df_datos_ordenados_vista)
        else:
            num_max_vista = int(input_num_str)

        if num_max_vista < 0: num_max_vista = 0
        num_max_vista = min(num_max_vista, len(df_datos_ordenados_vista))

    except ValueError:
        print("    Entrada inválida. No se generará nada.")
        input("    Presione Enter para continuar...")
        return

    if num_max_vista == 0:
        print("    Cantidad 0 seleccionada. No se generarán vistas rápidas.")
        input("    Presione Enter para continuar...")
        return

    df_final_para_vista = df_datos_ordenados_vista.head(num_max_vista)
    print(f"    Se generarán vistas rápidas para {len(df_final_para_vista)} expedientes.")

    # 5. Cargar PM.xlsx
    print(f"\n  4. Cargando PM desde: '{mode_config['pm_file_name']}'...")
    pm_set_actual_vista = cargar_pm_xlsx(mode_config) # Usas tu función existente

    # 6. Llamar a la función de lógica en GeneradorPredial_logica.py
    print(f"\n  5. Enviando {len(df_final_para_vista)} registros a la lógica de Predial para vista rápida...")
    GeneradorPredial_logica.generar_vista_rapida_dos_paginas_predial(
        df_datos_para_procesar_ordenados=df_final_para_vista,
        pm_set_actual=pm_set_actual_vista,
        config_predial_actual=mode_config,
        nombres_columnas_csv=nombres_col_csv,
        tipos_columnas_csv=tipos_col_csv
    )

    print(f"\n--- [UI] Proceso de Vista Rápida Finalizado ---")
    input("    Presione Enter para volver al menú de acciones...")

# generador.py

def sub_menu_acciones_modo(mode_config):
    """Muestra y maneja las acciones para un modo ya seleccionado."""
    while True:
        print(f"\n--- Opciones para Modo: {mode_config['mode_name']} ({mode_config['mode_type']}) ---")
        print("  ACCIONES DE GENERACIÓN:")
        print(f"  1. Generar Documento COMPLETO (solo nuevos)")
        print(f"  2. Generar SÓLO ÚLTIMA PÁGINA")
        print(f"  3. Generar RESTO del Documento")
        print(f"  4. Generar Expedientes ESPECÍFICOS")
        print("\n  ACCIONES DE ESCANEO:")
        if mode_config["mode_type"] == "PREDIAL":
            print(f"  5. PROCESAR DOCUMENTOS ESCANEADOS (PREDIAL)")
        else: # MULTAS
            print(f"  5. PROCESAR DOCUMENTOS ESCANEADOS (MULTAS)")
        print("\n  ACCIONES DE MANTENIMIENTO:")
        print(f"  A. Actualizar Registros en BD Maestra (Interactivo)")
        print(f"  U. Subir BD Maestra de este modo al Servidor")
        if mode_config["mode_type"] == "MULTAS":
             print(f"  R. Crear Reporte de Despachos (Maldonado Gallardo)")
             print(f"  F. Revisar Estado de Impresión de Flotillas")

        print("-" * 60)
        print(f"  V. Volver al menú principal de modos")

        opcion_accion = input(f"Seleccione una acción para '{mode_config['mode_name']}': ").strip().upper()

        if opcion_accion == 'V':
            break 

        modo_accion_solicitada = None
        
        if opcion_accion == '1':
            modo_accion_solicitada = mode_config["MODO_GENERACION_COMPLETO"]
        elif opcion_accion == '2':
            modo_accion_solicitada = mode_config["MODO_GENERACION_ULTIMA"]
        elif opcion_accion == '3':
            modo_accion_solicitada = mode_config["MODO_GENERACION_RESTO"]
        elif opcion_accion == '4':
            modo_accion_solicitada = mode_config["MODO_GENERACION_ESPECIFICOS"]
        elif opcion_accion == '5':
            if mode_config["mode_type"] == "MULTAS": 
                modo_accion_solicitada = MODO_ACCION_ESCANEAR_MULTAS
            elif mode_config["mode_type"] == "PREDIAL":
                modo_accion_solicitada = MODO_ACCION_ESCANEAR_PREDIAL
        elif opcion_accion == 'A':
            actualizar_bd_maestra_interactivo(mode_config)
            continue
        elif opcion_accion == 'U':
            print(f"\n  Subiendo Base de Datos Maestra para {mode_config['mode_name']}...")
            ruta_bd_maestra_local = Path(mode_config.get("master_db_file_path"))
            if ruta_bd_maestra_local.exists():
                subir_archivo_al_servidor(ruta_bd_maestra_local, mode_config)
                print("  -> Subida finalizada.")
            else:
                print(f"  (!) Error: No se encontró el archivo de BD Maestra en '{ruta_bd_maestra_local}'.")
            input("    Presione Enter para continuar...");
            continue
        elif opcion_accion == 'R' and mode_config["mode_type"] == "MULTAS":
            print(f"\n  Iniciando creación de Reporte de Despachos para {mode_config['mode_name']}...")
            GeneradorMultas_logica.crear_reporte_despachos_main(str(mode_config['base_path']), mode_config)
            input("    Presione Enter para continuar...");
            continue
        elif opcion_accion == 'F' and mode_config["mode_type"] == "MULTAS":
            print(f"\n  Iniciando revisión de estado para Flotillas en {mode_config['mode_name']}...")
            df_bd_maestra = cargar_bd_maestra_unificada(mode_config)
            GeneradorMultas_logica.crear_reporte_estado_flotillas(mode_config, df_bd_maestra)
            input("\n    Presione Enter para continuar...");
            continue
        else:
            print("Opción de acción inválida.")
        
        if modo_accion_solicitada:
            process_selected_mode_action(mode_config, modo_accion_solicitada)


def subir_archivo_al_servidor(ruta_local_archivo, config_modo):
    """
    Sube un archivo a una carpeta remota, creando la estructura de subdirectorios
    necesaria para que coincida con la ruta local.
    """
    print(f"    - Subiendo '{ruta_local_archivo.name}' al servidor...")
    
    # --- Lógica para construir la ruta remota completa ---
    ruta_base_servidor = Path("/srv/datos_gobierno")

    # 1. Obtiene la carpeta del modo, ej: "PREDIAL - AFC"
    try:
        remote_mode_folder = config_modo['base_path'].relative_to(Path(__file__).resolve().parent)
    except (ValueError, NameError):
        remote_mode_folder = Path(config_modo['base_path'].name)

    # 2. CALCULO CORREGIDO: Obtiene la ruta relativa del archivo DENTRO del modo.
    #    Ej: "DOCUMENTOS/COLONIAS/buenos_aires/archivo.pdf"
    try:
        sub_path_del_archivo = ruta_local_archivo.relative_to(config_modo['base_path'])
    except ValueError:
        sub_path_del_archivo = Path(ruta_local_archivo.name)
    
    # 3. Se construye la ruta final y completa que tendrá en el servidor.
    archivo_remoto_path = ruta_base_servidor / remote_mode_folder / sub_path_del_archivo
    ruta_remota_log = archivo_remoto_path.as_posix()
    print(f"      -> Destino: '{ruta_remota_log}'")

    try:
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect("asesorescloud.ddns.net", port=58123, username="afc", password="asesores", timeout=10)
        sftp = ssh.open_sftp()
        
        # 4. NUEVO: Asegurarse de que el directorio remoto exista, creándolo si es necesario.
        directorio_remoto_padre = archivo_remoto_path.parent
        
        partes_ruta = directorio_remoto_padre.as_posix().strip('/').split('/')
        ruta_acumulada = '/'
        for parte in partes_ruta:
            if not parte: continue
            ruta_acumulada += parte + '/'
            try:
                sftp.stat(ruta_acumulada)
            except FileNotFoundError:
                print(f"      - Creando directorio remoto: {ruta_acumulada}")
                sftp.mkdir(ruta_acumulada)

        # 5. Subir el archivo al destino final.
        sftp.put(str(ruta_local_archivo), archivo_remoto_path.as_posix())
        
        sftp.close()
        ssh.close()
        print(f"    -> Archivo subido exitosamente.")
        return True
    except Exception as e:
        print(f"    - (!) ERROR CRÍTICO al subir archivo al servidor: {e}")
        traceback.print_exc()
        return False
    
def sincronizar_todo_con_servidor():
    """
    Detecta todos los modos y sube la versión más reciente de su
    BASE_DE_DATOS.xlsx y base_datos_maestra_*.xlsx al servidor.
    """
    print("\n" + "="*30 + " INICIANDO SINCRONIZACIÓN MANUAL " + "="*30)
    modos_a_sincronizar = detectar_modos_disponibles()
    
    if not modos_a_sincronizar:
        print("No se encontraron modos válidos para sincronizar.")
        return

    print(f"Se sincronizarán {len(modos_a_sincronizar)} modos con el servidor.")
    
    for nombre_modo, config_modo in modos_a_sincronizar.items():
        print(f"\n--- Sincronizando Modo: {nombre_modo} ---")
        
        # 1. Sincronizar el archivo de la Base de Datos Maestra
        ruta_maestra_local = config_modo.get("master_db_file_path")
        if ruta_maestra_local and ruta_maestra_local.exists():
            print(f"  - Encontrado: {ruta_maestra_local.name}")
            subir_archivo_al_servidor(ruta_maestra_local, config_modo)
        else:
            print(f"  - (!) Advertencia: No se encontró el archivo de BD Maestra en '{ruta_maestra_local}'. No se pudo subir.")
            
        # 2. Sincronizar el archivo de Datos de Origen (BASE_DE_DATOS.xlsx)
        ruta_origen_local = config_modo.get("data_file_path")
        if ruta_origen_local and ruta_origen_local.exists():
            print(f"  - Encontrado: {ruta_origen_local.name}")
            subir_archivo_al_servidor(ruta_origen_local, config_modo)
        else:
            print(f"  - (!) Advertencia: No se encontró el archivo de Datos de Origen en '{ruta_origen_local}'. No se pudo subir.")
            
    print("\n" + "="*30 + " SINCRONIZACIÓN MANUAL COMPLETADA " + "="*30)



# --- Punto de Entrada del Script ---
if __name__ == '__main__':
    # Aquí deberías definir las funciones de utilidad que faltan (limpiar_texto, configurar_locale, etc.)
    # y las funciones _generar_documentos_predial_logic, _procesar_lote_multas_logic
    # y las auxiliares que estas últimas necesiten de GeneradorFER_logica.py y generadorMultas_logica.py

    # También las funciones del Paso 2 (cargar_bd_maestra_unificada, etc.) deben estar definidas ANTES de main_generador_loop.
    # Asegúrate de pegar las funciones del Paso 2 donde indiqué "# --- INICIO FUNCIONES PASO 2 ---"

    # Ejemplo de cómo llamarías a configurar_locale globalmente si la defines:
    # if not configurar_locale_universal():
    #     print("ADVERTENCIA: No se pudo configurar el locale correctamente.")

    main_generador_loop()